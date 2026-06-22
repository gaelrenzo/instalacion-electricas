#!/usr/bin/env python3
"""Exporta datos de Excel a formato LaTeX para el expediente."""
import openpyxl
from datetime import datetime, timedelta
from pathlib import Path

BASE = Path(__file__).resolve().parents[1] / "entregables"


def clean(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        epoch = datetime(1899, 12, 30)
        delta = val - epoch
        serial = delta.days + delta.seconds / 86400
        return f"{serial:.2f}"
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.2f}"
    return str(val).strip()


def fix_excel_number(val):
    """Reverses Excel's date-formatted numbers back to raw values."""
    if val is None:
        return None, ""
    if isinstance(val, datetime):
        epoch = datetime(1899, 12, 30)
        delta = val - epoch
        raw = delta.days + delta.seconds / 86400
        return raw, f"{raw:.2f}"
    if isinstance(val, (int, float)):
        return val, str(val)
    return val, str(val).strip()


def export_cuadro_cargas():
    wb = openpyxl.load_workbook(BASE / "calculos" / "cuadro-cargas.xlsx", data_only=True)
    ws = wb["Cuadro de Cargas"]
    lines = []
    for row in ws.iter_rows(min_row=5, max_row=10, values_only=True):
        items = [clean(v) for v in row]
        if items[0]:
            lines.append(" & ".join(items) + " \\\\")
    return "\n".join(lines)


def export_maxima_demanda():
    wb = openpyxl.load_workbook(BASE / "calculos" / "maxima-demanda.xlsx", data_only=True)
    ws = wb["Máxima Demanda"]
    lines = []
    for row in ws.iter_rows(min_row=5, max_row=21, values_only=True):
        items = [clean(v) for v in row]
        if items[0] and "TOTAL" not in str(items[1]).upper() and "NOTA" not in str(items[0]):
            lines.append(" & ".join(items[:7]) + " \\\\")
        elif items[0] and "TOTAL" in str(items[1]).upper():
            lines.append("\\midrule")
            lines.append("\\textbf{Total} & \\textbf{" + items[1] + "} & & & & & \\\\")
    return "\n".join(lines)


def export_metrados():
    wb = openpyxl.load_workbook(BASE / "presupuesto" / "metrados.xlsx", data_only=True)
    ws = wb["Metrados Eléctricos"]
    lines = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        items = [clean(v) for v in row]
        if items[0] or items[1]:
            lines.append(" & ".join(items[:6]) + " \\\\")
    return "\n".join(lines)


def export_presupuesto():
    wb = openpyxl.load_workbook(BASE / "presupuesto" / "presupuesto.xlsx", data_only=True)
    ws = wb["Presupuesto Eléctrico"]
    lines = []
    for row in ws.iter_rows(min_row=5, max_row=24, values_only=True):
        vals = list(row)
        items = []
        for v in vals:
            raw, formatted = fix_excel_number(v)
            items.append(formatted)
        if any(i for i in items):
            lines.append(" & ".join(items) + " \\\\")
    return "\n".join(lines)


def main():
    print("%%% CUADRO DE CARGAS %%%")
    print(export_cuadro_cargas())
    print()
    print("%%% MAXIMA DEMANDA %%%")
    print(export_maxima_demanda())
    print()
    print("%%% METRADOS %%%")
    print(export_metrados())
    print()
    print("%%% PRESUPUESTO %%%")
    print(export_presupuesto())


if __name__ == "__main__":
    main()
