#!/usr/bin/env python3
"""Genera tablas LaTeX desde los Excel del proyecto."""
import openpyxl
from datetime import datetime
from pathlib import Path

BASE = Path(__file__).resolve().parents[1] / "entregables"
CAP = Path(__file__).resolve().parents[1] / "expediente" / "capitulos"


def xl_num(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        epoch = datetime(1899, 12, 30)
        d = v - epoch
        raw = d.days + d.seconds / 86400
        return f"{raw:.2f}"
    if isinstance(v, float):
        return f"{int(v)}" if v == int(v) else f"{v:.2f}"
    return str(v).strip()


def read_cuadro():
    wb = openpyxl.load_workbook(BASE / "calculos" / "cuadro-cargas.xlsx", data_only=True)
    ws = wb["Cuadro de Cargas"]
    out = []
    for row in ws.iter_rows(min_row=5, max_row=10, values_only=True):
        cid, desc, pot, fd = [str(v).strip() if v else "" for v in row[:4]]
        cond, itm, idp = [str(v).strip() if v else "" for v in row[6:9]]
        if cid == "TOTAL":
            break
        if cid:
            out.append(f"{cid} & {desc} & {pot} & {fd} & {cond} & {itm} & {idp} \\\\")
    return "\n".join(out)


def read_demanda():
    wb = openpyxl.load_workbook(BASE / "calculos" / "maxima-demanda.xlsx", data_only=True)
    ws = wb["Máxima Demanda"]
    out = []
    for row in ws.iter_rows(min_row=5, max_row=21, values_only=True):
        item, desc, cant, pu, _, fd = [str(v).strip() if v else "" for v in row[:6]]
        if item and "TOTAL" not in desc.upper() and "NOTA" not in item.upper():
            out.append(f"{item} & {desc} & {cant} & {pu} & {fd} \\\\")
    return "\n".join(out)


def read_presupuesto():
    wb = openpyxl.load_workbook(BASE / "presupuesto" / "presupuesto.xlsx", data_only=True)
    ws = wb["Presupuesto Eléctrico"]
    out = []
    for row in ws.iter_rows(min_row=5, max_row=21, values_only=True):
        it, _, desc = [str(v).strip() if v else "" for v in row[:3]]
        und = str(row[3]).strip() if row[3] else ""
        cant_str = xl_num(row[4])
        pu_str = xl_num(row[5])
        if it:
            out.append(f"\\textbf{{{it}}} & {desc} & {und} & {cant_str} & {pu_str} \\\\")
    return "\n".join(out)


def read_metrados():
    wb = openpyxl.load_workbook(BASE / "presupuesto" / "metrados.xlsx", data_only=True)
    ws = wb["Metrados Eléctricos"]
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        it, cod, desc, und, cant, obs = [str(v).strip() if v else "" for v in row[:6]]
        if it or cod:
            out.append(f"{it} & {cod} & {desc} & {und} & {cant} & {obs} \\\\")
    return "\n".join(out)


def replace_between(content, start_marker, end_marker, new_content):
    """Replace text between two markers inclusive."""
    s = content.find(start_marker)
    if s < 0:
        return content
    e = content.find(end_marker, s)
    if e < 0:
        return content
    e += len(end_marker)
    return content[:s] + new_content + content[e:]


def main():
    cuadro_rows = read_cuadro()
    demanda_rows = read_demanda()
    pres_rows = read_presupuesto()
    meta_rows = read_metrados()

    # == 02-calculos-justificativos.tex ==
    calc = CAP / "02-calculos-justificativos.tex"
    txt = calc.read_text(encoding="utf-8")

    new_cuadro = (
        "\\section{Cuadro de cargas por circuito}\n\n"
        "\\begin{table}[H]\n\\centering\\small\\setlength{\\tabcolsep}{3pt}\n"
        "\\caption{Cuadro de cargas por circuito}\n"
        "\\label{tab:cuadro-cargas}\n"
        "\\begin{tabularx}{\\textwidth}{c Y c c Y Y Y}\n\\toprule\n"
        "\\textbf{Cto.} & \\textbf{Descripción} & \\textbf{Pot. (W)} & \\textbf{F.D.} & \\textbf{Conductor} & \\textbf{ITM} & \\textbf{ID} \\\\\n\\midrule\n"
        f"{cuadro_rows}\n"
        "\\midrule\n"
        "\\textbf{TOTAL} & \\textbf{Demanda consolidada} & & & Acometida: 2×10mm² Cu + PE & Gral: 2P-40A & ID: 2P-40A/30mA \\\\\n"
        "\\bottomrule\n\\end{tabularx}\n\\end{table}"
    )
    txt = replace_between(txt, "\\section{Cuadro de cargas por circuito}", "\\end{table}\n\n\\section{Levantamiento", new_cuadro + "\n\n\\section{Levantamiento")

    new_demanda = (
        "\\begin{table}[H]\n\\centering\\small\\setlength{\\tabcolsep}{3pt}\n"
        "\\caption{Levantamiento de cargas por ambiente}\n"
        "\\label{tab:levantamiento-cargas}\n"
        "\\begin{tabularx}{\\textwidth}{c Y c c c}\n\\toprule\n"
        "\\textbf{Item} & \\textbf{Ambiente / Carga} & \\textbf{Cant.} & \\textbf{Pot. Unit. (W)} & \\textbf{F. Demanda} \\\\\n\\midrule\n"
        f"{demanda_rows}\n"
        "\\bottomrule\n\\end{tabularx}\n\\end{table}"
    )
    txt = replace_between(txt, "\\section{Levantamiento de cargas por ambiente}", "\\end{table}\n\n\\section{Interpretacion", new_demanda + "\n\n\\section{Interpretacion")

    calc.write_text(txt, encoding="utf-8")
    print(f"✓ {calc.name} actualizado")

    # == 09-presupuesto.tex ==
    pres = CAP / "09-presupuesto.tex"
    txt2 = pres.read_text(encoding="utf-8")

    new_pres = (
        "\\section{Presupuesto general}\n\n"
        "\\begin{landscape}\n"
        "\\begin{table}[H]\n\\centering\\footnotesize\\setlength{\\tabcolsep}{4pt}\n"
        "\\caption{Presupuesto general estimado}\n"
        "\\label{tab:presupuesto-general}\n"
        "\\begin{tabularx}{\\textwidth}{c Y c c R{1.8cm}}\n\\toprule\n"
        "\\textbf{Item} & \\textbf{Descripción} & \\textbf{Und.} & \\textbf{Cant.} & \\textbf{P. Unit. (S/)} \\\\\n\\midrule\n"
        f"{pres_rows}\n"
        "\\bottomrule\n\\end{tabularx}\n\\end{table}\n"
        "\\end{landscape}\n\n"
        "\\section{Nota técnica}\n\n"
    )
    txt2 = replace_between(txt2, "\\section{Presupuesto general}", "\\section{Nota tecnica}", new_pres)
    pres.write_text(txt2, encoding="utf-8")
    print(f"✓ {pres.name} actualizado")

    # == 06-metrado.tex ==
    meta = CAP / "06-metrado.tex"
    txt3 = meta.read_text(encoding="utf-8")

    new_meta = (
        "\\section{Metrado general de materiales}\n\n"
        "\\begin{landscape}\n"
        "\\begin{table}[H]\n\\centering\\small\n"
        "\\caption{Metrado general de instalaciones eléctricas (datos Excel)}\n"
        "\\label{tab:metrado-excel}\n"
        "\\begin{tabularx}{\\textwidth}{c L{3.5cm} Y c c Y}\n\\toprule\n"
        "\\textbf{Item} & \\textbf{Código} & \\textbf{Descripción} & \\textbf{Und.} & \\textbf{Cant.} & \\textbf{Observación} \\\\\n\\midrule\n"
        f"{meta_rows}\n"
        "\\bottomrule\n\\end{tabularx}\n\\end{table}\n"
        "\\end{landscape}\n\n"
    )
    txt3 = replace_between(txt3, "\\section{Resumen de puntos electricos por circuito}", "\\section{Nota tecnica}", new_meta + "\\section{Nota tecnica}")
    # Add closing brace properly
    # Actually the replace replaces everything from Resumen to Nota. Let me handle better:
    meta.write_text(txt3, encoding="utf-8")
    print(f"✓ {meta.name} actualizado")


if __name__ == "__main__":
    main()
