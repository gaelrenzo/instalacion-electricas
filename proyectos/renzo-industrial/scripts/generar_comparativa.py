#!/usr/bin/env python3
"""Comparativa de cotizacion entre proveedores (Promart, Sodimac, Mercado Libre)
sobre el BOM de Renzo. Salidas en ``build/renzo-industrial/presupuesto`` y
copias a ``presupuesto/datos``.

Los precios base del BOM son referenciales de mercado; esta comparativa aplica
factores de margen por proveedor para estimar la dispersion de cotizaciones."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
BUILD = ROOT / "build" / "renzo-industrial" / "presupuesto"
DATOS = ROOT / "proyectos" / "renzo-industrial" / "presupuesto" / "datos"

# Factores de margen por proveedor y categoria (estimacion referencial).
# Promart/Sodimac: precio retail fisico; Mercado Libre: incluye flete variable.
FACTORES: dict[str, dict[str, float]] = {
    "Promart": {"default": 1.00, "cables": 1.05, "equipos de playa": 1.02, "grupo electrogeno": 1.00},
    "Sodimac": {"default": 0.98, "cables": 1.03, "equipos de playa": 1.00, "grupo electrogeno": 1.00},
    "Mercado Libre": {"default": 1.06, "cables": 1.08, "equipos de playa": 1.04, "grupo electrogeno": 1.02},
}

MONEDA_ICON = {"Promart": "Promart", "Sodimac": "Sodimac", "Mercado Libre": "Mercado Libre"}


def cargar(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def guardar(path: Path, data: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(data, encoding="utf-8")

def aplicar_factores(bom: dict) -> dict[str, dict[str, object]]:
    por_proveedor: dict[str, dict[str, object]] = {}
    for proveedor, margenes in FACTORES.items():
        total = 0.0
        por_categoria: dict[str, float] = {}
        n_items = 0
        for p in bom["materiales"]:
            factor = margenes.get(p["categoria"], margenes["default"])
            costo = round(p["costo_soles"] * factor, 2)
            total += costo
            por_categoria[p["categoria"]] = round(por_categoria.get(p["categoria"], 0.0) + costo, 2)
            n_items += 1
        por_proveedor[proveedor] = {"total_soles": round(total, 2), "por_categoria": por_categoria, "n_items": n_items}
    return por_proveedor


def principal() -> int:
    bom = cargar(BUILD / "bom_renzo.json")
    resultados = aplicar_factores(bom)
    base = bom["total_soles"]

    lineas = [
        "# Comparativa de proveedores - Renzo",
        "",
        f"BOM: **{bom['partidas']} partidas** - presupuesto base: **S/ {base:,.2f}**.",
        "",
        "| Proveedor | Total (S/) | Delta vs base |",
        "|---|---:|---:|",
    ]
    for proveedor, r in sorted(resultados.items(), key=lambda x: x[1]["total_soles"]):
        delta = (r["total_soles"] / base - 1) * 100
        lineas.append(f"| {MONEDA_ICON[proveedor]} | {r['total_soles']:,.2f} | {delta:+.1f}% |")
    lineas += ["", "> Factores de margen estimados (retail vs. flete). Solo referencia de dispersion, no cotizacion vinculante.", ""]

    guardar(BUILD / "comparativa-proveedores.md", "\n".join(lineas) + "\n")
    guardar(DATOS / "comparativa-proveedores.md", "\n".join(lineas) + "\n")

    # Excel de comparativa
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa"
        ws.append(["Proveedor", "Total (S/)", "Delta vs base (%)", "Partidas"])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for proveedor, r in sorted(resultados.items(), key=lambda x: x[1]["total_soles"]):
            ws.append([proveedor, r["total_soles"], round((r["total_soles"] / base - 1) * 100, 1), r["n_items"]])
        ws.append([])
        ws.append(["Presupuesto base", base, "", bom["partidas"]])
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10
        wb.save(BUILD / "cotizacion-comparativa.xlsx")
        wb.save(DATOS / "cotizacion-comparativa.xlsx")
    except ImportError:
        print("openpyxl no disponible; se omite Excel")

    print(json.dumps({"status": "PASS", "proveedores": {k: v["total_soles"] for k, v in resultados.items()}}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(principal())
