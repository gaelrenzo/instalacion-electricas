#!/usr/bin/env python3
"""Genera el BOM, metrados, presupuesto y comparativa de Renzo a partir de las
entradas canonicas (calculos, iluminacion, layout), sin modificar planos ni
expediente. Salidas en ``build/renzo-industrial/presupuesto`` y copias a
``presupuesto/datos``."""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[3]
PROJECT = ROOT / "proyectos/renzo-industrial"
BUILD = ROOT / "build" / "renzo-industrial" / "presupuesto"
DATOS = PROJECT / "presupuesto" / "datos"

# Precios unitarios referenciales (mercado peruano, 2026-07) en soles.
# Fuente: referencias de sodimac/promart/mercado libre (categoria homologada).
PRECIOS: dict[str, dict[str, Any]] = {
    "tuberia_pvc_20": {"descripcion": "Tuberia PVC SAP 20 mm (3 m) y accesorios", "und": "m", "pu": 3.50},
    "tuberia_pvc_25": {"descripcion": "Tuberia PVC SAP 25 mm (3 m) y accesorios", "und": "m", "pu": 4.80},
    "tuberia_pvc_32": {"descripcion": "Tuberia PVC SAP 32 mm (3 m) y accesorios", "und": "m", "pu": 6.50},
    "tuberia_pvc_40": {"descripcion": "Tuberia PVC SAP 40 mm (3 m) y accesorios", "und": "m", "pu": 9.00},
    "cable_2_5": {"descripcion": "Conductor de cobre XLPE/TW 2.5 mm2", "und": "m", "pu": 4.20},
    "cable_4": {"descripcion": "Conductor de cobre XLPE/TW 4 mm2", "und": "m", "pu": 5.80},
    "cable_10": {"descripcion": "Conductor de cobre XLPE/TW 10 mm2", "und": "m", "pu": 11.50},
    "cable_16": {"descripcion": "Conductor de cobre XLPE/TW 16 mm2", "und": "m", "pu": 15.00},
    "itm_2p_10": {"descripcion": "Interruptor termomagnetico 2P 10 A", "und": "und", "pu": 45.00},
    "itm_2p_16": {"descripcion": "Interruptor termomagnetico 2P 16 A", "und": "und", "pu": 52.00},
    "itm_2p_20": {"descripcion": "Interruptor termomagnetico 2P 20 A", "und": "und", "pu": 58.00},
    "itm_3p_10": {"descripcion": "Interruptor termomagnetico 3P 10 A", "und": "und", "pu": 68.00},
    "itm_3p_16": {"descripcion": "Interruptor termomagnetico 3P 16 A", "und": "und", "pu": 72.00},
    "itm_3p_40": {"descripcion": "Interruptor termomagnetico 3P 40 A", "und": "und", "pu": 128.00},
    "itm_3p_32": {"descripcion": "Interruptor termomagnetico 3P 32 A", "und": "und", "pu": 118.00},
    "itm_3p_20": {"descripcion": "Interruptor termomagnetico 3P 20 A", "und": "und", "pu": 95.00},
    "itm_4p_50": {"descripcion": "Interruptor termomagnetico 4P 50 A (general)", "und": "und", "pu": 320.00},
    "diferencial_30": {"descripcion": "Interruptor diferencial 30 mA", "und": "und", "pu": 95.00},
    "tablero": {"descripcion": "Tablero metalico con barra N y PE", "und": "und", "pu": 650.00},
    "luminaria_panel": {"descripcion": "Luminaria LED panel 36 W empotrada", "und": "und", "pu": 180.00},
    "luminaria_estanca": {"descripcion": "Luminaria LED estanca 18 W", "und": "und", "pu": 95.00},
    "luminaria_highbay": {"descripcion": "Luminaria LED industrial 50 W", "und": "und", "pu": 210.00},
    "proyector_100": {"descripcion": "Proyector LED exterior 100 W IP66", "und": "und", "pu": 250.00},
    "poste_80": {"descripcion": "Luminaria LED de poste 80 W", "und": "und", "pu": 320.00},
    "luminaria_emergencia": {"descripcion": "Luminaria de emergencia y senalizacion", "und": "und", "pu": 140.00},
    "tomacorriente": {"descripcion": "Tomacorriente doble con puesta a tierra", "und": "und", "pu": 55.00},
    "interruptor_pared": {"descripcion": "Interruptor de pared unipolar", "und": "und", "pu": 35.00},
    "caja_octogonal": {"descripcion": "Caja octogonal para luminaria", "und": "und", "pu": 8.00},
    "caja_rectangular": {"descripcion": "Caja rectangular para tomacorriente/interruptor", "und": "und", "pu": 7.00},
    "stp": {"descripcion": "Bomba sumergible de tanque 1.5 hp (STP)", "und": "und", "pu": 3200.00},
    "surtidor": {"descripcion": "Surtidor de doble manguera con cabeza electronica", "und": "und", "pu": 18000.00},
    "compresor": {"descripcion": "Compresor de aire de servicio 2.2 kW", "und": "und", "pu": 4800.00},
    "bomba_agua": {"descripcion": "Bomba de agua de servicio 1.5 kW", "und": "und", "pu": 1900.00},
    "bomba_fosa": {"descripcion": "Bomba de efluentes/fosa 1.5 kW", "und": "und", "pu": 1900.00},
    "grupo": {"descripcion": "Grupo electrogeno standby 37.5 kVA con ATS", "und": "und", "pu": 45000.00},
    "ups_fuel": {"descripcion": "UPS para combustible/control 1.5 kVA", "und": "und", "pu": 2400.00},
    "ups_it": {"descripcion": "UPS para POS/CCTV 2.0 kVA", "und": "und", "pu": 2400.00},
    "pozo_tierra": {"descripcion": "Pozo de puesta a tierra con varilla Cu 5/8x2.4 m", "und": "und", "pu": 1200.00},
    "pararrayo": {"descripcion": "Pararrayo h=12 m con radio 20 m", "und": "und", "pu": 6500.00},
    "cable_tierra_10": {"descripcion": "Conductor de tierra desnudo Cu 10 mm2", "und": "m", "pu": 9.00},
    "montaje": {"descripcion": "Montaje, pruebas y verificaciones", "und": "gbl", "pu": 15000.00},
}

CONDUCTOR_TUBERIA = {2.5: "tuberia_pvc_20", 4: "tuberia_pvc_25", 10: "tuberia_pvc_32", 16: "tuberia_pvc_40"}
CONDUCTOR_CABLE = {2.5: "cable_2_5", 4: "cable_4", 10: "cable_10", 16: "cable_16"}


def cargar(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def guardar(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def guardar_texto(path: Path, texto: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(texto, encoding="utf-8")


def circuito_con_pe(conductor: float) -> int:
    """Numero de conductores activos (fase/neutro) por circuito."""
    return 3 if conductor >= 4 else 2


def principal() -> int:
    calculos = cargar(BUILD.parent / "calculos" / "resumen-calculos.json")
    ilum = cargar(BUILD.parent / "calculos" / "iluminacion-resumen.json")
    layout = cargar(PROJECT / "arquitectura" / "datos" / "layout-grifo.json")
    if calculos.get("status") != "PASS":
        raise SystemExit("Calculos no en estado PASS")

    partidas: list[dict[str, Any]] = []
    circuitos = calculos["circuits"]

    for circuito in circuitos:
        ci = circuito["id"]
        conductor = float(circuito["conductor_mm2"])
        pe = float(circuito["pe_mm2"])
        longitud = float(circuito["length_m"])
        cable = CONDUCTOR_CABLE[conductor]
        tuberia = CONDUCTOR_TUBERIA[conductor]
        # Tuberia por circuito
        partidas.append({
            "codigo": f"{ci}-TUB",
            "descripcion": PRECIOS[tuberia]["descripcion"],
            "categoria": "ductos/tuberias",
            "und": "m",
            "cantidad": round(longitud, 1),
            "pu": PRECIOS[tuberia]["pu"],
            "circuito": ci,
            "fuente": f"circuito {ci}: longitud {longitud} m",
        })
        # Cable activo + neutro
        activos = circuito_con_pe(conductor)
        partidas.append({
            "codigo": f"{ci}-CAB",
            "descripcion": PRECIOS[cable]["descripcion"],
            "categoria": "cables",
            "und": "m",
            "cantidad": round(longitud * activos, 1),
            "pu": PRECIOS[cable]["pu"],
            "circuito": ci,
            "fuente": f"circuito {ci}: {activos} conductores activos x {longitud} m",
        })
        # Cable PE
        if pe > 0:
            pe_cable = CONDUCTOR_CABLE.get(pe, "cable_2_5")
            partidas.append({
                "codigo": f"{ci}-PE",
                "descripcion": f"{PRECIOS[pe_cable]['descripcion']} (PE)",
                "categoria": "cables",
                "und": "m",
                "cantidad": round(longitud, 1),
                "pu": PRECIOS[pe_cable]["pu"],
                "circuito": ci,
                "fuente": f"circuito {ci}: PE {pe} mm2 x {longitud} m",
            })
        # ITM
        if ci == "F-01":
            pass
        polo = "3" if circuito.get("breaker_poles") == 3 else "2"
        itm_key = f"itm_{polo}p_{int(float(circuito['breaker_a']))}"
        if itm_key not in PRECIOS:
            itm_key = f"itm_{polo}p_{int(float(circuito['breaker_a']))}"
        partidas.append({
            "codigo": f"{ci}-ITM",
            "descripcion": PRECIOS[itm_key]["descripcion"],
            "categoria": "interruptores termomagneticos",
            "und": "und",
            "cantidad": 1,
            "pu": PRECIOS[itm_key]["pu"],
            "circuito": ci,
            "fuente": f"cuadro de circuitos: ITM {int(float(circuito['breaker_a']))} A",
        })
        # Diferencial (todos los circuitos con RCD 30 mA)
        partidas.append({
            "codigo": f"{ci}-RCD",
            "descripcion": PRECIOS["diferencial_30"]["descripcion"],
            "categoria": "interruptores diferenciales",
            "und": "und",
            "cantidad": 1,
            "pu": PRECIOS["diferencial_30"]["pu"],
            "circuito": ci,
            "fuente": "cuadro de circuitos: RCD 30 mA",
        })

    # Alimentadores
    for alimentador in calculos["feeders"]:
        ai = alimentador["id"]
        fase = float(alimentador["phase_mm2"])
        neutro = float(alimentador["neutral_mm2"])
        pe = float(alimentador["pe_mm2"])
        longitud = float(alimentador["length_m"])
        cable_f = CONDUCTOR_CABLE.get(fase, "cable_16")
        cable_n = CONDUCTOR_CABLE.get(neutro, "cable_16")
        tuberia = CONDUCTOR_TUBERIA.get(fase, "tuberia_pvc_40")
        partidas.append({
            "codigo": f"{ai}-TUB",
            "descripcion": PRECIOS[tuberia]["descripcion"],
            "categoria": "ductos/tuberias",
            "und": "m",
            "cantidad": round(longitud, 1),
            "pu": PRECIOS[tuberia]["pu"],
            "circuito": ai,
            "fuente": f"alimentador {ai}: longitud {longitud} m",
        })
        # 3 fases + neutro
        partidas.append({
            "codigo": f"{ai}-CAB",
            "descripcion": PRECIOS[cable_f]["descripcion"] + " (3F)",
            "categoria": "cables",
            "und": "m",
            "cantidad": round(longitud * 3, 1),
            "pu": PRECIOS[cable_f]["pu"],
            "circuito": ai,
            "fuente": f"alimentador {ai}: 3 fases x {longitud} m",
        })
        partidas.append({
            "codigo": f"{ai}-NEU",
            "descripcion": PRECIOS[cable_n]["descripcion"] + " (N)",
            "categoria": "cables",
            "und": "m",
            "cantidad": round(longitud, 1),
            "pu": PRECIOS[cable_n]["pu"],
            "circuito": ai,
            "fuente": f"alimentador {ai}: neutro x {longitud} m",
        })
        partidas.append({
            "codigo": f"{ai}-PE",
            "descripcion": PRECIOS[CONDUCTOR_CABLE.get(pe, "cable_4")]["descripcion"] + " (PE)",
            "categoria": "cables",
            "und": "m",
            "cantidad": round(longitud, 1),
            "pu": PRECIOS[CONDUCTOR_CABLE.get(pe, "cable_4")]["pu"],
            "circuito": ai,
            "fuente": f"alimentador {ai}: PE {pe} mm2 x {longitud} m",
        })

    # Equipos y tableros (fuera de circuito)
    tableros = {"AL-TDE": 1, "AL-TDF": 1, "AL-TD-A1": 1}
    for key, cant in tableros.items():
        partidas.append({
            "codigo": f"{key}-TAB",
            "descripcion": PRECIOS["tablero"]["descripcion"] + f" ({key})",
            "categoria": "tableros",
            "und": "und",
            "cantidad": cant,
            "pu": PRECIOS["tablero"]["pu"],
            "circuito": key,
            "fuente": "cuadro de alimentadores",
        })
    partidas.append({"codigo": "GEN-ITM", "descripcion": PRECIOS["itm_4p_50"]["descripcion"], "categoria": "interruptores termomagneticos", "und": "und", "cantidad": 1, "pu": PRECIOS["itm_4p_50"]["pu"], "circuito": "GENERAL", "fuente": "interruptor general 50 A 4P"})

    # Luminarias segun memoria de iluminacion
    luminaria_mapa = {
        "ADM": "luminaria_panel", "OFI": "luminaria_panel", "SMAQ": "luminaria_highbay",
        "SS1": "luminaria_estanca", "SS2": "luminaria_estanca", "SS3": "luminaria_estanca",
        "VER": "luminaria_estanca", "DESP": "proyector_100", "PAT": "poste_80",
    }
    for ambiente in ilum["ambientes"]:
        key = luminaria_mapa.get(ambiente["id"])
        if key:
            partidas.append({
                "codigo": f"LUM-{ambiente['id']}",
                "descripcion": PRECIOS[key]["descripcion"],
                "categoria": "luminarias",
                "und": "und",
                "cantidad": int(ambiente["n_luminarias"]),
                "pu": PRECIOS[key]["pu"],
                "circuito": "ILUMINACION",
                "fuente": f"memoria de iluminacion: {ambiente['nombre']} ({ambiente['n_luminarias']} lum)",
            })

    # Alumbrado de emergencia y senalizacion
    partidas.append({"codigo": "EMG-LUM", "descripcion": PRECIOS["luminaria_emergencia"]["descripcion"], "categoria": "luminarias", "und": "und", "cantidad": 6, "pu": PRECIOS["luminaria_emergencia"]["pu"], "circuito": "S-03", "fuente": "EM.010 art. 11.1"})

    # Tomacorrientes (layout: 5 tomas en IE-02 + criticos)
    partidas.append({"codigo": "TC-IE02", "descripcion": PRECIOS["tomacorriente"]["descripcion"], "categoria": "tomacorrientes", "und": "und", "cantidad": 5, "pu": PRECIOS["tomacorriente"]["pu"], "circuito": "A1-03", "fuente": "lamina IE-02 (5 tomas)"})
    partidas.append({"codigo": "INT-PARED", "descripcion": PRECIOS["interruptor_pared"]["descripcion"], "categoria": "interruptores", "und": "und", "cantidad": 8, "pu": PRECIOS["interruptor_pared"]["pu"], "circuito": "A1-01", "fuente": "lamina IE-02"})
    cajas_oct = sum(int(a["n_luminarias"]) for a in ilum["ambientes"]) + 6
    partidas.append({"codigo": "CAJA-OCT", "descripcion": PRECIOS["caja_octogonal"]["descripcion"], "categoria": "cajas", "und": "und", "cantidad": cajas_oct, "pu": PRECIOS["caja_octogonal"]["pu"], "circuito": "GENERAL", "fuente": "una caja por luminaria + emergencia"})
    partidas.append({"codigo": "CAJA-REC", "descripcion": PRECIOS["caja_rectangular"]["descripcion"], "categoria": "cajas", "und": "und", "cantidad": 13, "pu": PRECIOS["caja_rectangular"]["pu"], "circuito": "GENERAL", "fuente": "tomas + interruptores"})

    # Equipos de playa observados en DWG
    partidas.append({"codigo": "EQ-STP", "descripcion": PRECIOS["stp"]["descripcion"], "categoria": "equipos de playa", "und": "und", "cantidad": 3, "pu": PRECIOS["stp"]["pu"], "circuito": "F-01..F-03", "fuente": "DWG: TK-1..TK-3"})
    partidas.append({"codigo": "EQ-SURT", "descripcion": PRECIOS["surtidor"]["descripcion"], "categoria": "equipos de playa", "und": "und", "cantidad": 2, "pu": PRECIOS["surtidor"]["pu"], "circuito": "F-04..F-05", "fuente": "DWG: islas 1 y 2"})
    partidas.append({"codigo": "EQ-CAIRE", "descripcion": PRECIOS["compresor"]["descripcion"], "categoria": "equipos de servicio", "und": "und", "cantidad": 1, "pu": PRECIOS["compresor"]["pu"], "circuito": "F-07", "fuente": "sala de maquinas"})
    partidas.append({"codigo": "EQ-BAGUA", "descripcion": PRECIOS["bomba_agua"]["descripcion"], "categoria": "equipos de servicio", "und": "und", "cantidad": 1, "pu": PRECIOS["bomba_agua"]["pu"], "circuito": "F-08", "fuente": "sala de maquinas"})
    partidas.append({"codigo": "EQ-BFOSA", "descripcion": PRECIOS["bomba_fosa"]["descripcion"], "categoria": "equipos de servicio", "und": "und", "cantidad": 1, "pu": PRECIOS["bomba_fosa"]["pu"], "circuito": "F-09", "fuente": "fosa de agua"})
    partidas.append({"codigo": "EQ-GRUPO", "descripcion": PRECIOS["grupo"]["descripcion"], "categoria": "grupo electrogeno", "und": "und", "cantidad": 1, "pu": PRECIOS["grupo"]["pu"], "circuito": "EMERGENCIA", "fuente": f"Cummins C30D6 {calculos['generator']['selected_nameplate_kva']} kVA"})
    partidas.append({"codigo": "EQ-UPS-FUEL", "descripcion": PRECIOS["ups_fuel"]["descripcion"], "categoria": "ups", "und": "und", "cantidad": 1, "pu": PRECIOS["ups_fuel"]["pu"], "circuito": "F-04..F-06", "fuente": "UPS-FUEL"})
    partidas.append({"codigo": "EQ-UPS-IT", "descripcion": PRECIOS["ups_it"]["descripcion"], "categoria": "ups", "und": "und", "cantidad": 1, "pu": PRECIOS["ups_it"]["pu"], "circuito": "S-01", "fuente": "UPS-IT"})

    # Puesta a tierra
    for num, pos in enumerate(("PAT", "PAT2"), 1):
        partidas.append({"codigo": f"TIERRA-{num}", "descripcion": PRECIOS["pozo_tierra"]["descripcion"] + f" ({pos})", "categoria": "puesta a tierra", "und": "und", "cantidad": 1, "pu": PRECIOS["pozo_tierra"]["pu"], "circuito": "GENERAL", "fuente": f"DWG: {pos}"})
    partidas.append({"codigo": "TIERRA-CABLE", "descripcion": PRECIOS["cable_tierra_10"]["descripcion"], "categoria": "puesta a tierra", "und": "m", "cantidad": 60, "pu": PRECIOS["cable_tierra_10"]["pu"], "circuito": "GENERAL", "fuente": "malla equipotencial IE-04"})
    partidas.append({"codigo": "RAYO", "descripcion": PRECIOS["pararrayo"]["descripcion"], "categoria": "puesta a tierra", "und": "und", "cantidad": 1, "pu": PRECIOS["pararrayo"]["pu"], "circuito": "GENERAL", "fuente": "DWG: h=12 m R=20 m"})

    # Montaje
    partidas.append({"codigo": "MONT-JE", "descripcion": PRECIOS["montaje"]["descripcion"], "categoria": "montaje", "und": "gbl", "cantidad": 1, "pu": PRECIOS["montaje"]["pu"], "circuito": "GENERAL", "fuente": "instalacion, telurometro, ensayos"})

    for p in partidas:
        p["costo_soles"] = round(p["cantidad"] * p["pu"], 2)

    total = round(sum(p["costo_soles"] for p in partidas), 2)
    por_categoria: dict[str, float] = {}
    for p in partidas:
        por_categoria[p["categoria"]] = round(por_categoria.get(p["categoria"], 0.0) + p["costo_soles"], 2)

    bom = {
        "schema_version": 1,
        "proyecto": "renzo-industrial",
        "generado": date.today().isoformat(),
        "fuentes": {
            "calculos": str((BUILD.parent / "calculos" / "resumen-calculos.json").relative_to(ROOT)),
            "iluminacion": str((BUILD.parent / "calculos" / "iluminacion-resumen.json").relative_to(ROOT)),
            "layout": "proyectos/renzo-industrial/arquitectura/datos/layout-grifo.json",
        },
        "partidas": len(partidas),
        "total_soles": total,
        "por_categoria": por_categoria,
        "materiales": partidas,
    }
    guardar(BUILD / "bom_renzo.json", bom)
    guardar(DATOS / "bom_renzo.json", bom)

    # Markdown de metrados
    lineas = [f"# Metrados y presupuesto - Renzo", "", f"Partidas: **{len(partidas)}**.", f"Total referencial: **S/ {total:,.2f}**.", "", "| Codigo | Descripcion | Categoria | Und | Cant | P.U. (S/) | Parcial (S/) | Circuito |", "|---|---|---:|---:|---:|---:|---:|---|"]
    for p in partidas:
        lineas.append(f"| {p['codigo']} | {p['descripcion']} | {p['categoria']} | {p['und']} | {p['cantidad']} | {p['pu']:,.2f} | {p['costo_soles']:,.2f} | {p['circuito']} |")
    lineas.append("")
    lineas.append("## Total por categoria")
    lineas.append("")
    lineas.append("| Categoria | Total (S/) |")
    lineas.append("|---:|---:|")
    for cat, total_cat in sorted(por_categoria.items(), key=lambda x: -x[1]):
        lineas.append(f"| {cat} | {total_cat:,.2f} |")
    guardar_texto(BUILD / "metrados-renzo.md", "\n".join(lineas) + "\n")
    guardar_texto(DATOS / "metrados-renzo.md", "\n".join(lineas) + "\n")

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "Metrados"
        headers = ["Codigo", "Descripcion", "Categoria", "Und", "Cantidad", "P.U. (S/)", "Parcial (S/)", "Circuito", "Fuente"]
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for p in partidas:
            ws.append([p["codigo"], p["descripcion"], p["categoria"], p["und"], p["cantidad"], p["pu"], p["costo_soles"], p["circuito"], p["fuente"]])
        ws.append([])
        ws.append(["TOTAL", "", "", "", "", "", total, "", ""])
        ws.cell(row=ws.max_row, column=7).font = Font(bold=True)
        for col in ("A", "B", "C", "D", "E", "F", "G", "H", "I"):
            ws.column_dimensions[col].width = 14 if col == "A" else (45 if col == "B" else (22 if col == "C" else 10))
        wb.save(BUILD / "metrados-renzo.xlsx")
        wb.save(DATOS / "metrados-renzo.xlsx")
    except ImportError:
        print("openpyxl no disponible; se omite Excel")

    print(json.dumps({"status": "PASS", "partidas": len(partidas), "total_soles": total, "output": str(BUILD)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(principal())
